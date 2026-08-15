"""
Demo pack generator — synthetic Whitmore Family Office statements.

Regenerates all five demo documents in demo_pack/ from scratch. Safe to
re-run any number of times (idempotent — each run simply overwrites the
prior output with byte-for-byte identical content given the same source).

Usage:
    python3 demo_pack/generate.py
"""

from __future__ import annotations

import os

from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    BaseDocTemplate,
    Frame,
    NextPageTemplate,
    PageTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))

SYNTHETIC_NOTICE = (
    "This document is SYNTHETIC SAMPLE DATA generated for demonstration purposes only. "
    "It does not represent a real account, holding, or transaction."
)

# ---------------------------------------------------------------------------
# PDF helpers
# ---------------------------------------------------------------------------

NAVY = colors.HexColor("#1c2b3a")
STEEL = colors.HexColor("#4a5a6a")
HAIRLINE = colors.HexColor("#c7cdd3")
PALE = colors.HexColor("#eef1f4")


def _footer(canvas, doc, statement_ref: str, issuer: str):
    canvas.saveState()
    canvas.setStrokeColor(HAIRLINE)
    canvas.setLineWidth(0.5)
    canvas.line(0.75 * inch, 0.72 * inch, LETTER[0] - 0.75 * inch, 0.72 * inch)
    canvas.setFont("Helvetica", 7.2)
    canvas.setFillColor(STEEL)
    canvas.drawString(0.75 * inch, 0.56 * inch, f"{issuer}  |  Ref: {statement_ref}")
    canvas.drawRightString(
        LETTER[0] - 0.75 * inch, 0.56 * inch, f"Page {doc.page}"
    )
    canvas.setFont("Helvetica-Oblique", 6.6)
    canvas.setFillColor(colors.HexColor("#9aa5af"))
    canvas.drawCentredString(LETTER[0] / 2, 0.40 * inch, SYNTHETIC_NOTICE)
    canvas.restoreState()


def build_pdf(filename: str, statement_ref: str, issuer: str, flow: list):
    path = os.path.join(HERE, filename)
    doc = BaseDocTemplate(
        path,
        pagesize=LETTER,
        leftMargin=0.75 * inch,
        rightMargin=0.75 * inch,
        topMargin=0.65 * inch,
        bottomMargin=0.95 * inch,
        title=filename,
    )
    frame = Frame(
        doc.leftMargin,
        doc.bottomMargin,
        doc.width,
        doc.height,
        id="normal",
    )

    def on_page(canvas, d):
        _footer(canvas, d, statement_ref, issuer)

    template = PageTemplate(id="main", frames=[frame], onPage=on_page)
    doc.addPageTemplates([template])
    doc.build(flow)
    return path


def styles():
    ss = getSampleStyleSheet()
    ss.add(
        ParagraphStyle(
            "Issuer",
            parent=ss["Normal"],
            fontName="Helvetica-Bold",
            fontSize=13,
            textColor=NAVY,
            spaceAfter=2,
        )
    )
    ss.add(
        ParagraphStyle(
            "DocTitle",
            parent=ss["Normal"],
            fontName="Helvetica-Bold",
            fontSize=10.5,
            textColor=STEEL,
            spaceAfter=8,
        )
    )
    ss.add(
        ParagraphStyle(
            "Meta",
            parent=ss["Normal"],
            fontName="Helvetica",
            fontSize=9,
            textColor=colors.HexColor("#333c45"),
            leading=13,
            spaceAfter=2,
        )
    )
    ss.add(
        ParagraphStyle(
            "Section",
            parent=ss["Normal"],
            fontName="Helvetica-Bold",
            fontSize=10,
            textColor=NAVY,
            spaceBefore=14,
            spaceAfter=6,
        )
    )
    ss.add(
        ParagraphStyle(
            "Body",
            parent=ss["Normal"],
            fontName="Helvetica",
            fontSize=8.6,
            leading=12.4,
            textColor=colors.HexColor("#20272e"),
        )
    )
    ss.add(
        ParagraphStyle(
            "Disclaimer",
            parent=ss["Normal"],
            fontName="Helvetica-Oblique",
            fontSize=7.4,
            leading=10.4,
            textColor=STEEL,
            spaceBefore=10,
        )
    )
    return ss


def data_table(rows, col_widths, header=True, align_right_from=1):
    t = Table(rows, colWidths=col_widths, hAlign="LEFT")
    style = [
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.4),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LINEBELOW", (0, 0), (-1, -1), 0.4, HAIRLINE),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]
    if align_right_from is not None:
        style.append(("ALIGN", (align_right_from, 0), (-1, -1), "RIGHT"))
    if header:
        style += [
            ("BACKGROUND", (0, 0), (-1, 0), NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("LINEBELOW", (0, 0), (-1, 0), 0.6, NAVY),
        ]
    t.setStyle(TableStyle(style))
    return t


# ---------------------------------------------------------------------------
# Document 1 — Ridgeline Capital Partners VI capital statement
# ---------------------------------------------------------------------------

def build_doc1():
    ss = styles()
    flow = []
    flow.append(Paragraph("RIDGELINE CAPITAL MANAGEMENT LLC", ss["Issuer"]))
    flow.append(
        Paragraph(
            "Ridgeline Capital Partners VI, L.P. &mdash; Quarterly Capital Account Statement",
            ss["DocTitle"],
        )
    )
    flow.append(
        Paragraph(
            "Reporting Period Ended: June 30, 2026  |  Statement Issued: July 18, 2026",
            ss["Meta"],
        )
    )
    flow.append(
        Paragraph("Prepared for: Whitmore Family Office (LP Investor ID: WFO-2231)", ss["Meta"])
    )
    flow.append(Paragraph("General Partner: Ridgeline Capital Management LLC", ss["Meta"]))
    flow.append(
        Paragraph(
            "Vintage: 2022  |  Strategy: Growth Equity  |  Geography: North America",
            ss["Meta"],
        )
    )
    flow.append(
        Paragraph(
            "Fund Term: 10 years + two 1-year extensions (expected wind-down 2032&ndash;2034)",
            ss["Meta"],
        )
    )

    flow.append(Paragraph("Capital Account Summary (USD)", ss["Section"]))
    rows = [
        ["", "Since Inception"],
        ["Total Commitment", "$8,000,000"],
        ["Paid-In Capital (Called)", "$5,200,000"],
        ["Cumulative Distributions", "$340,000"],
        ["Reported Net Asset Value (NAV)", "$6,100,000"],
        ["Remaining Unfunded Commitment", "$2,800,000"],
    ]
    flow.append(data_table(rows, [3.2 * inch, 2.3 * inch]))
    flow.append(Spacer(1, 6))
    flow.append(
        Paragraph(
            "Note on Capital Calls. The General Partner retains the right to call remaining "
            "unfunded capital for follow-on investments and fund expenses on ten (10) "
            "business days&rsquo; written notice. Failure to fund a capital call constitutes an "
            "Event of Default under Section 6.3 of the Limited Partnership Agreement and may "
            "result in forfeiture of up to 100% of the defaulting Partner&rsquo;s interest in the Fund.",
            ss["Body"],
        )
    )

    flow.append(Paragraph("Schedule of Investments &mdash; Top Holdings (unaudited)", ss["Section"]))
    rows2 = [
        ["Portfolio Company", "Sector", "Fair Value", "% of Fund NAV"],
        ["Aurex Data Centres, Inc.", "Digital Infrastructure", "$854,000", "14.0%"],
        ["Vertexa Software Inc", "Enterprise SaaS", "$671,000", "11.0%"],
        ["Brightline Energy Storage", "Grid & Storage", "$549,000", "9.0%"],
        ["Cobalt Freight Systems", "Logistics Tech", "$427,000", "7.0%"],
        ["Other (12 companies)", "Various", "$3,599,000", "59.0%"],
    ]
    flow.append(data_table(rows2, [1.9 * inch, 1.7 * inch, 1.1 * inch, 0.9 * inch]))

    flow.append(
        Paragraph(
            "Valuations reflect the most recent round of financing or, where unavailable, comparable "
            "public market multiples as of the period end date. LP interests are illiquid; no secondary "
            "market exists and transfers require GP consent. Figures are unaudited. "
            "&copy; Ridgeline Capital Management LLC &mdash; Confidential. Not for redistribution.",
            ss["Disclaimer"],
        )
    )

    return build_pdf(
        "01_ridgeline_capital_VI_capital_statement.pdf",
        "RCP6-Q2-2026-WFO2231",
        "Ridgeline Capital Management LLC",
        flow,
    )


# ---------------------------------------------------------------------------
# Document 2 — Calder Marine equity purchase agreement / holding confirmation
# ---------------------------------------------------------------------------

def build_doc2():
    ss = styles()
    flow = []
    flow.append(Paragraph("IRONVALE PARTNERS LLP", ss["Issuer"]))
    flow.append(
        Paragraph(
            "Equity Purchase Confirmation &amp; Holding Statement &mdash; Private Placement",
            ss["DocTitle"],
        )
    )
    flow.append(Paragraph("Transaction / Valuation Date: July 20, 2026", ss["Meta"]))
    flow.append(Paragraph("Prepared for: Whitmore Family Office", ss["Meta"]))
    flow.append(
        Paragraph("Arranger / Placement Agent: Ironvale Partners LLP", ss["Meta"])
    )
    flow.append(Paragraph("Confirmation No.: IVP-EPA-20260720-014", ss["Meta"]))

    flow.append(Paragraph("Transaction Summary", ss["Section"]))
    rows = [
        ["Target Entity", "Calder Marine Terminals, LLC"],
        ["Interest Acquired", "40% Membership Interest"],
        ["Purchase Consideration", "$3,100,000"],
        ["Closing Date", "July 20, 2026"],
        ["Counterparty / Arranger", "Ironvale Partners LLP (acting as arranger)"],
        ["Seller", "Calder Founders Holdco LLC"],
    ]
    flow.append(data_table(rows, [2.6 * inch, 3.0 * inch], align_right_from=None))

    flow.append(Paragraph("Business Description", ss["Section"]))
    flow.append(
        Paragraph(
            "Calder Marine Terminals, LLC operates two bulk marine terminals handling dry-bulk and "
            "breakbulk cargo. The Company is privately held; Whitmore Family Office&rsquo;s 40% "
            "membership interest was acquired directly from existing members with Ironvale Partners "
            "LLP acting as transaction arranger. This holding is illiquid: no established secondary "
            "market exists for membership interests in the Company, transfer is subject to a right of "
            "first refusal held by the remaining members, and no near-term liquidity event is "
            "contemplated as of the date of this confirmation.",
            ss["Body"],
        )
    )

    flow.append(Paragraph("Governance & Reporting", ss["Section"]))
    flow.append(
        Paragraph(
            "Whitmore Family Office holds board observer rights but no controlling interest. Quarterly "
            "unaudited financial statements are to be delivered by the Company within 45 days of "
            "quarter end; the initial delivery is expected with the September 30, 2026 period.",
            ss["Body"],
        )
    )

    flow.append(
        Paragraph(
            "This confirmation is provided for informational purposes and does not constitute investment, "
            "legal, or tax advice. Valuation reflects transaction price at closing and is not marked to "
            "any subsequent market event. &copy; Ironvale Partners LLP &mdash; Confidential.",
            ss["Disclaimer"],
        )
    )

    return build_pdf(
        "02_calder_marine_equity_purchase_agreement.pdf",
        "IVP-EPA-20260720-014",
        "Ironvale Partners LLP",
        flow,
    )


# ---------------------------------------------------------------------------
# Document 5 — Gladstone Industrial Park appraisal summary
# ---------------------------------------------------------------------------

def build_doc5():
    ss = styles()
    flow = []
    flow.append(Paragraph("HARLOW & VANCE APPRAISAL GROUP", ss["Issuer"]))
    flow.append(Paragraph("Summary Appraisal Report &mdash; Industrial Property", ss["DocTitle"]))
    flow.append(Paragraph("Effective Date of Value: June 15, 2026", ss["Meta"]))
    flow.append(Paragraph("Prepared for: Whitmore Family Office", ss["Meta"]))
    flow.append(Paragraph("Appraiser of Record: Harlow & Vance Appraisal Group", ss["Meta"]))
    flow.append(Paragraph("Report No.: HV-2026-0615-GIP", ss["Meta"]))

    flow.append(Paragraph("Property Identification", ss["Section"]))
    rows = [
        ["Property", "Gladstone Industrial Park, Units 4–9"],
        ["Location", "Savannah, GA"],
        ["Property Type", "Multi-tenant Industrial / Flex Warehouse"],
        ["Owning Entity", "Gladstone Park Holdings LLC (100% Whitmore Family Office)"],
        ["Prior Appraisal on File", "March 2025"],
    ]
    flow.append(data_table(rows, [2.2 * inch, 3.4 * inch], align_right_from=None))

    flow.append(Paragraph("Valuation Conclusion (USD)", ss["Section"]))
    rows2 = [
        ["Reconciled Market Value (As-Is)", "$9,800,000"],
        ["First Mortgage Outstanding &mdash; Coastal Trust Bank", "$5,200,000"],
        ["Mortgage Rate / Maturity", "6.10% fixed, maturing August 2032"],
        ["Net Equity", "$4,600,000"],
    ]
    flow.append(data_table(rows2, [3.6 * inch, 2.0 * inch]))

    flow.append(Paragraph("Marketability & Liquidity", ss["Section"]))
    flow.append(
        Paragraph(
            "The subject is considered illiquid relative to publicly traded securities. Based on "
            "comparable listings and absorption in the Savannah industrial submarket, a marketing "
            "period of six to nine months would be required to achieve the concluded value in an "
            "arm&rsquo;s-length sale. This appraisal supersedes the prior report on file dated March 2025; "
            "material market movement and one new lease since that date support the revised conclusion.",
            ss["Body"],
        )
    )

    flow.append(Paragraph("Scope & Methodology", ss["Section"]))
    flow.append(
        Paragraph(
            "Value concluded via reconciliation of the Sales Comparison and Income Capitalization "
            "Approaches. The Cost Approach was considered but given diminished weight due to the "
            "age of improvements. This is a summary report intended to comply with USPAP Standards "
            "Rule 2-2(b) and is not a self-contained appraisal report.",
            ss["Body"],
        )
    )

    flow.append(
        Paragraph(
            "This report is prepared solely for the use of Whitmore Family Office and its lender of "
            "record and may not be relied upon by any other party. &copy; Harlow &amp; Vance Appraisal "
            "Group &mdash; Confidential.",
            ss["Disclaimer"],
        )
    )

    return build_pdf(
        "05_gladstone_industrial_appraisal.pdf",
        "HV-2026-0615-GIP",
        "Harlow & Vance Appraisal Group",
        flow,
    )


# ---------------------------------------------------------------------------
# XLSX helpers
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1C2B3A")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
TITLE_FONT = Font(bold=True, size=13, color="1C2B3A", name="Calibri")
SUBTITLE_FONT = Font(size=10, color="4A5A6A", name="Calibri")
META_FONT = Font(size=9.5, color="20272E", name="Calibri")
BODY_FONT = Font(size=10, name="Calibri")
NOTE_FONT = Font(size=9, italic=True, color="4A5A6A", name="Calibri")
THIN = Side(style="thin", color="C7CDD3")
BORDER = Border(bottom=THIN)


def _sheet_header(ws, title_rows, start_row=1):
    r = start_row
    for text, font in title_rows:
        cell = ws.cell(row=r, column=1, value=text)
        cell.font = font
        r += 1
    return r + 1


def _write_table(ws, start_row, headers, data_rows, col_widths):
    hr = start_row
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=hr, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = BORDER
    for ri, row in enumerate(data_rows, start=hr + 1):
        for ci, val in enumerate(row, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = BODY_FONT
            c.border = BORDER
            if ci > 1 and isinstance(val, (int, float)):
                c.number_format = "#,##0"
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return hr + len(data_rows) + 1


def _synthetic_footer(ws, row, max_col):
    cell = ws.cell(row=row + 1, column=1, value=SYNTHETIC_NOTICE)
    cell.font = Font(size=8, italic=True, color="9AA5AF")
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=max_col)


# ---------------------------------------------------------------------------
# Document 3 — Ironvale Private Credit schedule (two sheets)
# ---------------------------------------------------------------------------

def build_doc3():
    path = os.path.join(HERE, "03_ironvale_private_credit_schedule.xlsx")
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Loan Schedule"
    r = _sheet_header(
        ws1,
        [
            ("Ironvale Private Credit", TITLE_FONT),
            ("Direct Lending Schedule — Loan Level Detail", SUBTITLE_FONT),
            ("As of: July 31, 2026", META_FONT),
            ("Prepared for: Whitmore Family Office", META_FONT),
            ("Statement Ref: IVPC-2026-07-LS-002", META_FONT),
        ],
    )
    headers = [
        "Facility ID",
        "Borrower",
        "Principal Outstanding (USD)",
        "Coupon",
        "Maturity",
        "Security",
        "Covenant Status",
    ]
    rows = [
        [
            "IVP-2026-003",
            "Vertexa Software, Inc.",
            2400000,
            "12.50%",
            "2029-09-30",
            "Unsecured",
            "Compliant",
        ],
        [
            "IVP-2025-021",
            "Brightline Energy Storage Inc.",
            1650000,
            "10.25%",
            "2028-06-30",
            "Senior Secured",
            "Compliant",
        ],
    ]
    last_row = _write_table(ws1, r, headers, rows, [16, 30, 22, 10, 13, 16, 16])
    _synthetic_footer(ws1, last_row, len(headers))

    ws2 = wb.create_sheet("Notes")
    r2 = _sheet_header(
        ws2,
        [
            ("Ironvale Private Credit", TITLE_FONT),
            ("Facility Notes & Credit Support", SUBTITLE_FONT),
            ("As of: July 31, 2026", META_FONT),
            ("Prepared for: Whitmore Family Office", META_FONT),
        ],
    )
    note_lines = [
        "Facility IVP-2026-003 (Vertexa Software, Inc.):",
        "This facility is unsecured and, in the event of a liquidation or insolvency proceeding "
        "involving the Borrower, ranks behind all secured creditors of the Borrower with respect to "
        "recovery of principal and accrued interest.",
        "",
        "Credit support: this facility carries a sponsor guarantee provided by Ridgeline Capital "
        "Partners VI, L.P., covering the full outstanding principal balance of the facility. The "
        "guarantee is unsecured and unfunded; it represents a contractual undertaking of the "
        "guarantor rather than a pledge of specific collateral.",
        "",
        "Facility IVP-2025-021 (Brightline Energy Storage Inc.):",
        "Senior secured by a first-priority lien over substantially all assets of the Borrower. No "
        "third-party guarantee is attached to this facility.",
    ]
    rr = r2
    for line in note_lines:
        c = ws2.cell(row=rr, column=1, value=line)
        c.font = BODY_FONT if line else BODY_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws2.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=6)
        rr += 1
    ws2.column_dimensions["A"].width = 100
    _synthetic_footer(ws2, rr, 6)

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Document 4 — Brightwater custody statement
# ---------------------------------------------------------------------------

def build_doc4():
    path = os.path.join(HERE, "04_brightwater_custody_statement.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Reserve Portfolio"

    r = _sheet_header(
        ws,
        [
            ("Brightwater Trust Company", TITLE_FONT),
            ("Custody & Brokerage Statement", SUBTITLE_FONT),
            ("As of: August 5, 2026", META_FONT),
            ("Account: Whitmore Family Office — Reserve Portfolio", META_FONT),
            ("Account No.: BW-WFO-40021", META_FONT),
        ],
    )
    headers = ["Asset", "Asset Type", "Market Value (USD)", "Settlement / Liquidity"]
    rows = [
        ["Cash & Money Market", "Cash", 1200000, "T+0"],
        ["US Large Cap Equity Index", "Public Equity", 2400000, "T+2"],
        ["Municipal Bond Ladder", "Fixed Income", 900000, "T+2"],
    ]
    last_row = _write_table(ws, r, headers, rows, [28, 18, 20, 22])

    total_row = last_row + 1
    ws.cell(row=total_row, column=1, value="Total").font = Font(bold=True, size=10)
    ws.cell(row=total_row, column=3, value=4500000).font = Font(bold=True, size=10)
    ws.cell(row=total_row, column=3).number_format = "#,##0"
    ws.cell(row=total_row, column=3).border = BORDER
    ws.cell(row=total_row, column=1).border = BORDER

    note = ws.cell(
        row=total_row + 2,
        column=1,
        value=(
            "Settlement/liquidity reflects standard settlement cycle for the asset class under normal "
            "market conditions and does not account for market-impact costs on large liquidation orders."
        ),
    )
    note.font = NOTE_FONT
    note.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=total_row + 2, start_column=1, end_row=total_row + 2, end_column=4)

    _synthetic_footer(ws, total_row + 3, 4)
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    os.makedirs(HERE, exist_ok=True)
    paths = [
        build_doc1(),
        build_doc2(),
        build_doc3(),
        build_doc4(),
        build_doc5(),
    ]
    for p in paths:
        size = os.path.getsize(p)
        print(f"{os.path.basename(p):55s} {size:>8,d} bytes")


if __name__ == "__main__":
    main()
